from openpyxl import workbook, load_workbook
from Properties_and_Capacities import Section_Properties as SP
import pandas as pd

class Excel_Print:

    def __init__(self, xlName, xlsheetname):
        
        self.xlFileName = xlName
        self.xlSheetName = xlsheetname
        

    def get_max_row(self):

        wb = load_workbook(self.xlFileName)

        #Create new sheet if xlsheetname is not in the workbook
        if self.xlSheetName not in wb.sheetnames:
            wb.create_sheet(self.xlSheetName)
        
        wb.active = wb[self.xlSheetName]
        ws = wb.active

        return ws.max_row

    # Printing the elastic section properties to Excel
    def Elastic_Prop_to_Excel(self, Geom_Input, n = 8):

        BmTable, BmModulus, BmInertia = SP.Calc_Elastic_Prop(Geom_Input, n)
    
        with pd.ExcelWriter(self.xlFileName, engine='openpyxl', mode='a',if_sheet_exists="overlay") as writer:
            BmTable.to_excel(writer, sheet_name = self.xlSheetName, startrow = self.get_max_row() + 2, index=False, startcol=2)
    
        Total = BmTable.sum()

        wb = load_workbook(self.xlFileName)
        wb.active = wb[self.xlSheetName]
        ws = wb.active
        sum_row = ws.max_row + 1

        ws["k"+str(sum_row)]= BmInertia
        ws["f"+str(sum_row)]= Total["Area"]
        ws["h"+str(sum_row)]= Total["AY"]
        wb.save(self.xlFileName)
        wb.close()

        with pd.ExcelWriter(self.xlFileName, engine='openpyxl', mode='a',if_sheet_exists="overlay") as writer:
        
            BmModulus.to_excel(writer, sheet_name = self.xlSheetName, startrow = self.get_max_row() + 2, startcol = 5,header=True, index=False)
