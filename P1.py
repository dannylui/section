import pandas as pd

from openpyxl import workbook, load_workbook

# Calculates the elastic section property of a I Section
def Elastic_Section_Prop(Section: pd.DataFrame, n: int = 8):
 
    y = []
    mod_ratio_n = []

    if n == 0:
        Section = Section.drop([0]).reset_index(drop=True)
        # print(Section["Width"] )
                      

    # Calculates y 
    for i in range(0,len(Section['Width'])):
        if i == 0:
            y.append(Section.loc[i,"Thickness/Height"]/2)
            if n == 0:
                mod_ratio_n.append(1)
            else:
                mod_ratio_n.append(n)
            
        else:
            y.append(Section.loc[i,"Thickness/Height"]/2 + y[i-1] + Section.loc[i-1,"Thickness/Height"]/2)
            mod_ratio_n.append(1)


    Section["Area"] = Section["Width"] * Section ["Thickness/Height"] / mod_ratio_n

    Section["Y"] = y
    Section["AY"] = Section["Area"] * Section ["Y"] 

    # print(Section["Area"])
    # print(Section["Width"] * Section ["Thickness/Height"])

    Total = Section.sum()
    Centroid = Total['AY'] / Total['Area']

    Section["D"] = Section["Y"] - Centroid
    Section["AD^2"] = Section["Area"] * Section["D"]**2

    Section["I0"] = Section['Width'] * Section["Thickness/Height"]**3 / 12 / mod_ratio_n
    Total = Section.sum()

    Inertia =  Total["AD^2"] + Total["I0"]

    BM_height = Total["Thickness/Height"]

    if Section.loc[0,"Element"] == 'Slab':
        BM_height -= Section.loc[0,'Thickness/Height']
      
    Y_BOT = Total["Thickness/Height"] - Centroid
    Y_TOP = BM_height - Y_BOT
    
    S_TOP = Inertia / Y_TOP
    S_BOT = Inertia / Y_BOT

    Modulus = pd.DataFrame({'Location':["Top","Bottom"],'Distance y':[Y_TOP,Y_BOT],'Section Modulus':[S_TOP,S_BOT ]})

    # print(Total)

    # print(Section)
    # print(Modulus)
    # print(Inertia)
    
    return Section, Modulus, Inertia

# Get the last row in the excel
def get_max_row(xlName, xlsheetname):

    wb = load_workbook(xlName)

    #Create new sheet if xlsheetname is not in the workbook
    if xlsheetname not in wb.sheetnames:
        wb.create_sheet(xlsheetname)
    
    wb.active = wb[xlsheetname]
    ws = wb.active

    return ws.max_row

# Printing the elastic section properties to Excel
def Elastic_Prop_to_Excel(xlName, xlsheetname, Section, n = 8):

    BmTable, BmModulus, BmInertia = Elastic_Section_Prop(Section, n)
   
    with pd.ExcelWriter(xlFilename, engine='openpyxl', mode='a',if_sheet_exists="overlay") as writer:
        BmTable.to_excel(writer, sheet_name = xlsheetname, startrow = get_max_row(xlName, xlsheetname) + 2, index=False, startcol=2)
   
    Total = BmTable.sum()
    # sum_row = ws.max_row+10


    wb = load_workbook(xlName)
    wb.active = wb[xlsheetname]
    ws = wb.active
    sum_row = ws.max_row+1

    ws["k"+str(sum_row)]= BmInertia
    ws["f"+str(sum_row)]= Total["Area"]
    ws["h"+str(sum_row)]= Total["AY"]
    wb.save(xlFilename)
    wb.close()

    with pd.ExcelWriter(xlFilename, engine='openpyxl', mode='a',if_sheet_exists="overlay") as writer:
    
        BmModulus.to_excel(writer, sheet_name = xlsheetname, startrow = get_max_row(xlName, xlsheetname) + 2, startcol = 5,header=True, index=False)

#AASHTO 6.10.2.1
def web_proportion_chk(Geom_Input):
    if Geom_Input['Long stiffener'] == 'no':
        return Geom_Input['D_web'] / Geom_Input['t_web'] <= 150
    else:
        return Geom_Input['D_web'] / Geom_Input['t_web'] <= 300



#Calculating Vp
def Plastic_Shear_Vp(Geom_Input: dict):
    return 0.58 * Geom_Input['fyw'] * Geom_Input['D_web'] * Geom_Input['t_web']

# returns true if stiffened, false if not. Assume no longitudinal stiffener
def is_stiffened(Geom_Input: dict, Stiff_input: dict):
    if Stiff_input['Panel'] == 'Interior' and Stiff_input['Spacing d0'] <= Geom_Input['D_web'] * 3:
        print ("Check stiffened capacity using 6.10.9.3")
        return True
    elif Stiff_input['Panel'] == 'End' and Stiff_input['Spacing d0'] <= Geom_Input['D_web'] * 1.5:
        print ("Check stiffened capacity using 6.10.9.3.3")
        return True
    else:
        print ("Check unstiffened capacity using 6.10.9.2")
        return False


def Shear_buckling_k (Geom_Input, Stiff_input):
    if not is_stiffened(Geom_Input, Stiff_input):
        print("k value per AASHTO 6.10.9.2")
        return 5
    else:
        print("Calculate k per AASHTO 6.10.9.3.2-7")
        return 5 + 5 / (Stiff_input['Spacing d0'] / Geom_Input['D_web']) ** 2

# Calculate Shear-buckling resistance to the shear yield Strength
    #AASHTO 6.10.9.3.2
def ratio_C(Geom_Input, Stiff_input):
    D = Geom_Input['D_web']
    tw = Geom_Input['t_web']
    E = Geom_Input['E']
    Fyw = Geom_Input['fyw']
    k = Shear_buckling_k(Geom_Input, Stiff_input)
    print("k=" + str(k))

    if D/tw <= 1.12 * (E * k / Fyw) ** 0.5:
        print("Calculate C per AASHTO 6.10.9.3.2-4")
        return 1
    elif D/tw < 1.4* (E * k / Fyw) ** 0.5:
        print("Calculate C per AASHTO 6.10.9.3.2-5")
        return 1.12 / (D / tw) * (E * k / Fyw) ** 0.5
    else:
        print("Calculate C per AASHTO 6.10.9.3.2-6")
        return 1.57 / ((D / tw) ** 2) * (E * k / Fyw)
    

def capacity_Vn (Geom_Input, Stiff_input):
    D = Geom_Input['D_web']
    tw = Geom_Input['t_web']
    ttf = Geom_Input['t_tf']
    btf = Geom_Input['b_tf']
    tbf = Geom_Input['t_bf']
    bbf = Geom_Input['b_bf']

    C = ratio_C(Geom_Input, Stiff_input)
    print("C=" + str(C))

    if Stiff_input['Panel'] == 'Interior':
        if (2 * D * tw)/(btf*ttf + bbf * tbf) <= 2.5:             #AASHTO 6.10.9.3.2-1
            print("Calculate Vn per AASHTO 6.10.9.3.2-1")
            return Plastic_Shear_Vp(Geom_Input) * (C + 0.87 * (1-C)/(1+(Stiff_input['Spacing d0'] / Geom_Input['D_web'])**2)**0.5)
        else:  #AASHTO 6.10.9.3.2-8
           print("Calculate Vn per AASHTO AASHTO 6.10.9.3.2-8")
           return Plastic_Shear_Vp(Geom_Input) * (C + 0.87 * (1-C)/((1+(Stiff_input['Spacing d0'] / Geom_Input['D_web'])**2)**0.5 + (Stiff_input['Spacing d0'] / Geom_Input['D_web'])))
    else: 
        return Plastic_Shear_Vp(Geom_Input) * C

def check_Vu(Vu, Vn):
    if Vu > Vn:
        return "NG"
    else:
        return "OK"

# Calculating the plastic neutral axis for positive flexure
    # Ref AASHTO Table D6.1-1
# conservative ignore rebars for now - to be implemented later *****
def PNA (Geom_Input: dict):
    Ps = 0.85 * Geom_Input['fc'] *Geom_Input['b_slab']*Geom_Input[ 't_slab']        # Slab
    Pt = Geom_Input['fy_bf'] * Geom_Input['b_bf'] * Geom_Input['t_bf']              # Bottom Flange/ tension
    Pc = Geom_Input['fy_tf'] * Geom_Input['b_tf'] * Geom_Input['t_tf']              # Top Flange/ compression
    Pw = Geom_Input['fyw'] * Geom_Input['D_web'] * Geom_Input['t_web']              # Web

    Y_bar = 0
    Mp = 0
    PNA_Loc = ''

    if Pt + Pw >= Pc + Ps:      # in web
        print("PNA is in the web")
        PNA_Loc = 'Web'
        Y_bar = Geom_Input['D_web']/2 * ((Pt-Pc-Ps)/Pw+1)                   # PNA from the top of the web
     
        ds = Y_bar + Geom_Input['t_tf'] + Geom_Input['t_haunch'] + Geom_Input[ 't_slab'] / 2         # Distance from PNA to center of the slab 
        dc = Y_bar + Geom_Input['t_tf'] / 2                 # Distance from PNA to center of the compression flange
        dt = Geom_Input['D_web'] - Y_bar + Geom_Input['t_bf'] / 2         # Distance from PNA to center of the compression flange

        Mp = Pw/(2 * Geom_Input['D_web']) * (Y_bar**2 + (Geom_Input['D_web']-Y_bar)**2) + (Ps*ds + Pc*dc + Pt * dt)
        
        print("ds = " + str(ds))
        print("dw = " + str(dw))
        print("dt = " + str(dt))
        
        print("Ybar = " + str(Y_bar))
        print("Mp = " + str(Mp))

    elif Pt + Pw + Pc >= Ps:    # in the top flange
        print("PNA is in the top flange")
        PNA_Loc = 'Top Flange'
        Y_bar = Geom_Input['t_tf']/2 * ((Pw + Pt - Ps)/Pc + 1)  

        ds = Y_bar + Geom_Input['t_haunch'] + Geom_Input['t_slab'] / 2         # Distance from PNA to center of the slab 
        dw = Geom_Input['t_tf'] - Y_bar + Geom_Input['D_web'] / 2           # Distance from PNA to center of the web 
        dt = Geom_Input['t_tf'] - Y_bar + Geom_Input['D_web'] + Geom_Input['t_bf'] / 2         # Distance from PNA to center of the compression flange
        
        Mp = Pc/(2 * Geom_Input['t_tf']) * (Y_bar**2 + (Geom_Input['t_tf']-Y_bar)**2) + (Ps*ds + Pw*dw + Pt * dt)
        
        print("ds = " + str(ds))
        print("dw = " + str(dw))
        print("dt = " + str(dt))

        print("Ybar = " + str(Y_bar))
        print("Mp = " + str(Mp))

    else:
        print("PNA is in the deck")
        PNA_Loc = 'Slab'

        Y_bar = Geom_Input[ 't_slab']  * (Pw + Pt + Pc)/Ps

        dc = Geom_Input['t_slab'] - Y_bar + Geom_Input['t_haunch'] + Geom_Input['t_tf'] / 2                 # Distance from PNA to center of the compression flange
        dw = Geom_Input['t_slab'] - Y_bar + Geom_Input['t_haunch'] + Geom_Input['t_tf'] + Geom_Input['D_web'] / 2           # Distance from PNA to center of the web 
        dt = Geom_Input['t_slab'] - Y_bar + Geom_Input['t_haunch'] + Geom_Input['t_tf'] + Geom_Input['D_web'] + Geom_Input['t_bf'] / 2         # Distance from PNA to center of the compression flange
        Mp = (Y_bar**2 * Ps)/ (2 * Geom_Input['t_slab']) + (Pc*dc + Pw*dw + Pt * dt)
        
        print("ds = " + str(ds))
        print("dw = " + str(dw))
        print("dt = " + str(dt))

        print("Ybar = " + str(Y_bar))
        print("Mp = " + str(Mp))

    return Y_bar, Mp, PNA_Loc

# AASHTO 6.10.6.2.2
def check_compactness(Geom_Input):
    
    if Geom_Input['fy_tf'] <= 70 and Geom_Input['fy_bf'] <= 70:
        print ("Flange strength is ok")
    else:
        print ("Flange strength is NG")

    if web_proportion_chk(Geom_Input):
        print('Web check per 6.10.2.1 is OK')
    else:
        print('Web check per 6.10.2.1 is NG')
    
    Y_bar, _ , PNA_loc = PNA(Geom_Input)
    Dcp = 0
    
    if PNA_loc == "Web":
        Dcp = Y_bar

    if 2 * Dcp/Geom_Input['t_web'] <= 3.76 * (Geom_Input['E']/Geom_Input['fy_tf'])**0.5:
        print('Web slenderness check is OK')
    else:
        print('Web slenderness check is NG')
    return  None

# n = modular ratio, eta = importance ratio
# AASHTO D6.2.2
def yield_moment (Geom_Input, Forces_input, n, eta):

    Section = pd.DataFrame({'Element':["Slab","Top Flange","Web","Bottom Flange"],
                       'Width': [Geom_Input['b_slab'], Geom_Input['b_tf'], Geom_Input['t_web'], Geom_Input['b_bf']],
                       'Thickness/Height':[Geom_Input['t_slab'],Geom_Input['t_tf'],Geom_Input['D_web'],Geom_Input['t_bf']]})
    
    _,S_NC,_ = Elastic_Section_Prop(Section, 0)
    _,S_LT,_ = Elastic_Section_Prop(Section, 3*n)
    _,S_ST,_ = Elastic_Section_Prop(Section, n)

 # Bottom Flange    
    S_NC_bf = S_NC['Section Modulus'][1]
    S_LT_bf = S_LT['Section Modulus'][1]
    S_ST_bf = S_ST['Section Modulus'][1]

    MAD_bf = (Geom_Input['fy_bf'] / eta - 1.25 * Forces_input['M_dc1']/S_NC_bf - (1.25 * Forces_input['M_dc2'] + 1.5 * Forces_input['M_dw']) / S_LT_bf) * S_ST_bf

    My_bf = 1.25 * Forces_input['M_dc1'] + 1.25 * Forces_input['M_dc2'] + 1.5 * Forces_input['M_dw'] + MAD_bf

 # Top Flange
    S_NC_tf = S_NC['Section Modulus'][0]
    S_LT_tf = S_LT['Section Modulus'][0]
    S_ST_tf = S_ST['Section Modulus'][0]

    MAD_tf = (Geom_Input['fy_tf'] / eta - 1.25 * Forces_input['M_dc1']/S_NC_tf - (1.25 * Forces_input['M_dc2'] + 1.5 * Forces_input['M_dw']) / S_LT_tf) * S_ST_tf

    My_tf = 1.25 * Forces_input['M_dc1'] + 1.25 * Forces_input['M_dc2'] + 1.5 * Forces_input['M_dw'] + MAD_tf

    return min(My_tf, My_bf)

# AASHTO 6.5.4.2
Resist_factors_phi = {'phi_v':1.0,    # Shear
                      'phi_f':1.0 }   # Flexure

Stiffener_Input = {'Panel':'Interior',      # End/ Interior only
                   'Spacing d0': 300}       # Spacing in inches
  

# Input materials and geometry
Input = {'fyw':50, 'fy_tf':50, 'fy_bf':50, 'E':29000,
         'fc':4,
         'Long stiffener':'no', 'Trans stiffener': 'yes',      # Should be yes/no only 
         'b_slab':114, 't_slab':9, 't_haunch':4, 
         'b_tf':16, 't_tf':1.0, 
         'D_web':69, 't_web':0.5, 
         'b_bf':18, 't_bf':1.75}

# Forces at section from analysis
Forces = {'M_dc1': 2202*12, 'M_dc2': 335*12,'M_dw': 322*12,'M_LL': 12222}

modular_ratio_n = 8       # modular ratio, Es/Ec

Import_factor = 1     # importance factor, essential, etc

BmSect = pd.DataFrame({'Element':["Slab","Top Flange","Web","Bottom Flange"],
                       'Width': [Input['b_slab'], Input['b_tf'], Input['t_web'], Input['b_bf']],
                       'Thickness/Height':[Input['t_slab'],Input['t_tf'],Input['D_web'],Input['t_bf']]})

#-----------------------

# xlFilename = "SectionProp.xlsx"
# xlSheet = "Sectio3"

# Elastic_Prop_to_Excel(xlFilename, xlSheet, BmSect, 0)
# Prop_to_Excel(xlFilename, xlSheet, BmSect, 8)
# Prop_to_Excel(xlFilename, xlSheet, BmSect, 24)

#----------------------
# print(Plastic_Shear_Vp(Input))

# print(is_stiffened(Input, Stiffener_Input))
# print("k = " + str(Shear_buckling_k(Input, Stiffener_Input)))

# print("C=" + str(ratio_C(Input, Stiffener_Input)))
# print("Vn=" + str(capacity_Vn(Input, Stiffener_Input)))

#--------------------

# print(PNA(Input))

# print(yield_moment(Input, Forces, modular_ratio_n, Import_factor)/12)

check_compactness(Input)

# BmTable.to_excel(xlFilename, index=False, startcol=2)

# wb = load_workbook(xlFilename)
# ws = wb.active

# ws["k"+str(ws.max_row+1)]= BmInertia

# wb.save(xlFilename)

# with pd.ExcelWriter(xlFilename, engine='openpyxl', mode='a',if_sheet_exists="overlay") as writer:
    
#     BmModulus.to_excel(writer, startrow = ws.max_row + 5, startcol = 5,header=True, index=False)

