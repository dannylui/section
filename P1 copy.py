import pandas as pd

from openpyxl import workbook, load_workbook

# def hello(name, age=10):
#     return "something", "else"

# x, y = hello("a")


BmSect = pd.DataFrame({'Element':["Slab","Top Flange","Web","Bottom Flange"],'Width':[50,12,0.75,12],'Thickness/Height':[8,0.5,24,0.5]})

def SectionProp(Section, n = 8):
 
    y = []
    nRat = []

    # Calculates y 
    for i in range(0,len(Section['Width'])):
        if i == 0:
            y.append(Section.loc[i,"Thickness/Height"]/2)
            nRat.append(n)
            
        else:
            y.append(Section.loc[i,"Thickness/Height"]/2 + y[i-1] + Section.loc[i-1,"Thickness/Height"]/2)
            nRat.append(1)


    Section["Area"] = Section["Width"] * Section ["Thickness/Height"] / nRat

    Section["Y"] = y
    Section["AY"] = Section["Area"] * Section ["Y"] 

    print(Section["Area"])
    print(Section["Width"] * Section ["Thickness/Height"])

    Total = Section.sum()
    Centroid = Total['AY'] / Total['Y']

    Section["D"] = Section["Y"] - Centroid
    Section["AD^2"] = Section["Area"] * Section["D"]**2

    Section["I0"] = Section['Width'] * Section["Thickness/Height"]**3 / 12 / nRat
    Total = Section.sum()

    Inertia =  Total["AD^2"] + Total["I0"]

    BM_height = Total["Thickness/Height"]

    Y_TOP = Centroid
    Y_BOT = BM_height - Centroid

    S_TOP = Inertia / Y_TOP
    S_BOT = Inertia / Y_BOT

    Modulus = pd.DataFrame({'Location':["Top","Bottom"],'Distance y':[Y_TOP,Y_BOT],'Section Modulus':[S_TOP,S_BOT ]})

    print(Total)

    print(Section)
    print(Modulus)
    print(Inertia)
    
    return Section, Modulus, Inertia


def get_max_row(xlName, xlsheetname):
    wb = load_workbook(xlName)

    #Create new sheet if xlsheetname is not in the workbook
    if xlsheetname not in wb.sheetnames:
        wb.create_sheet(xlsheetname)
    
    wb.active = wb[xlsheetname]
    ws = wb.active
    return ws.max_row



def ProptoExcel(xlName, xlsheetname, Section, n = 8):

    BmTable, BmModulus, BmInertia = SectionProp(Section, n)
   
    # wb = load_workbook(xlName)

    # #Create new sheet if xlsheetname is not in the workbook
    # if xlsheetname not in wb.sheetnames:
    #     wb.create_sheet(xlsheetname)
    
    # wb.active = wb[xlsheetname]
    # ws = wb.active

    # # wb.save(xlFilename)

    with pd.ExcelWriter(xlFilename, engine='openpyxl', mode='a',if_sheet_exists="overlay") as writer:
        BmTable.to_excel(writer, sheet_name = xlsheetname, startrow = ws.max_row + 2, index=False, startcol=2)
        sum_row = ws.max_row+2
   
    Total = BmTable.sum()
    # sum_row = ws.max_row+10

    ws["k"+str(sum_row)]= BmInertia
    ws["f"+str(sum_row)]= Total["Area"]
    ws["h"+str(sum_row)]= Total["AY"]
    wb.save(xlFilename)

    with pd.ExcelWriter(xlFilename, engine='openpyxl', mode='a',if_sheet_exists="overlay") as writer:
    
        BmModulus.to_excel(writer, sheet_name = xlsheetname, startrow = ws.max_row + 2, startcol = 5,header=True, index=False)

        

xlFilename = "SectionProp.xlsx"
xlSheet = "Section7"

ProptoExcel(xlFilename, xlSheet, BmSect, 8)
# ProptoExcel(xlFilename, xlSheet, BmSect, 24)

# BmTable.to_excel(xlFilename, index=False, startcol=2)

# wb = load_workbook(xlFilename)
# ws = wb.active

# ws["k"+str(ws.max_row+1)]= BmInertia

# wb.save(xlFilename)

# with pd.ExcelWriter(xlFilename, engine='openpyxl', mode='a',if_sheet_exists="overlay") as writer:
    
#     BmModulus.to_excel(writer, startrow = ws.max_row + 5, startcol = 5,header=True, index=False)

