from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

ws["B2"] = "Location"
ws["c2"] = "Width"
ws["d2"] = "Thickness"
ws["e2"] = "Area"
ws["f2"] = "Y"
ws["g2"] = "AY"
ws["h2"] = "Io"
ws["i2"] = "Ad^2"


ws["B3"] = "Top Flange"
ws["c3"] = 12
ws["d3"] = 0.5

ws["b4"] = "Web"
ws["c4"] = 0.75
ws["d4"] = 24

ws["B5"] = "Bottom Flange"
ws["c5"] = 12
ws["d5"] = 0.5

Y = 0

for row in range(3, 6):
    # Area
    ws["e" + str(row)] = f"=D{row} * C{row}"

    # Y from top
    ws["f" + str(row)] = f"=D{row} / 2 + {Y}"
    Y = Y + ws["d" + str(row)].value

    # A*Y
    ws["g" + str(row)] = f"=e{row} * f{row}"

    # Moment of Inertia
    ws["h" + str(row)] = f"=c{row} * d{row}^3/12"

for col in range(5, 10):
    char = get_column_letter(col)

    ws[char+"6"]=f"={char} 3 + {char} " 

wb.save("SectionProp.xlsx")
